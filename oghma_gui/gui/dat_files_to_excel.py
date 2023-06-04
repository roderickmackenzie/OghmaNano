# -*- coding: utf-8 -*-
#
#   OghmaNano - Organic and hybrid Material Nano Simulation tool
#   Copyright (C) 2008-2022 Roderick C. I. MacKenzie r.c.i.mackenzie at googlemail.com
#
#   https://www.oghma-nano.com
#
#   Permission is hereby granted, free of charge, to any person obtaining a
#   copy of this software and associated documentation files (the "Software"),
#   to deal in the Software without restriction, including without limitation
#   the rights to use, copy, modify, merge, publish, distribute, sublicense, 
#   and/or sell copies of the Software, and to permit persons to whom the
#   Software is furnished to do so, subject to the following conditions:
#
#   The above copyright notice and this permission notice shall be included
#   in all copies or substantial portions of the Software.
#
#   THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS
#   OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#   FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
#   THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#   LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#   OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE 
#   SOFTWARE.
#

## @package dat_files_to_excel
#  Load and dump a dat file into a dat class
#

from str2bool import str2bool
from bytes2str import str2bytes
from bytes2str import bytes2str

def dat_files_to_excel(output_file,dat_files):
	try:
		from openpyxl import Workbook
		from openpyxl.chart import Reference, Series, ScatterChart, LineChart
		#from openpyxl.compat import range
	except:
		print("openpyxl not found")

	if output_file.endswith(".xlsx")==False:
		output_file=output_file+".xlsx"

	wb = Workbook()
	ws_data = wb.active
	#ws_data = wb.create_sheet(title="data")
	data_set=0

	#chart
	c1 = LineChart()
	c1.title = bytes2str(dat_files[0].title)
	#c1.style = 13
	c1.height=20
	c1.width=20

	for data in dat_files:
		start_col=data_set*2+1
		ws_data.cell(column=start_col, row=1, value=bytes2str(data.y_label)+" ("+bytes2str(data.y_units)+") ")
		ws_data.cell(column=start_col+1, row=1, value=bytes2str(data.data_label)+" ("+bytes2str(data.data_units)+") ")

		row_pos=0
		for i in range(0,data.y_len):
			row_pos=i+2
			ws_data.cell(column=start_col, row=row_pos, value=data.y_scale[i])
			ws_data.cell(column=start_col+1, row=row_pos, value=data.data[0][0][i])

		data = Reference(ws_data, min_col=start_col, max_col=start_col+1, min_row=1, max_row=row_pos)
		c1.add_data(data, titles_from_data=True)

		data_set=data_set+1


	c1.y_axis.title = bytes2str(dat_files[0].data_label)+" ("+bytes2str(dat_files[0].data_units)+") "
	c1.x_axis.title = bytes2str(dat_files[0].y_label)+" ("+bytes2str(dat_files[0].y_units)+") "


	ws_data.add_chart(c1, "G4")


	#print("about to save1")
	try:
		wb.save(filename = output_file)
	except:
		return False

	return
	max=data[0].y_len
	for d in data:
		if d.y_len>max:
			max=d.y_len
	out=[]
	line=""
	for i in range(0,len(data)):
		line=bytes2str(line+data[i].key_text)+" "+bytes2str(data[i].y_label)+","+bytes2str(data[i].key_text)+" "+bytes2str(data[i].data_label)+","

	line=line[:-1]
	out.append(line)

	for i in range(0,max):
		line=""
		for ii in range(0,len(data)):
			y_text=""
			data_text=""
			if i<data[ii].y_len:
				y_text=str('{:.8e}'.format(float(data[ii].y_scale[i])))
				data_text=str('{:.8e}'.format(float(data[ii].data[0][0][i])))
			line=line+y_text+","+data_text+","
		line=line[:-1]
		out.append(line)

	inp_save_lines_to_file(file_name,out)


